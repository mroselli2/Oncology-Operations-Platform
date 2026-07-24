#!/usr/bin/env python3
"""
SUPERSEDED by oncology_pipeline/ -- kept for reference only. Do not run this
file as part of normal operation; its logic (pathway/milestone generation,
the five agent functions, the six forced_scenarios) has been migrated into
oncology_pipeline/generation/patients.py and oncology_pipeline/agents/.

Synthetic Oncology Scheduling Optimization Agent
================================================
Fully synthetic data only. No real PHI.

Implements a lightweight multi-agent style workflow for:
1. Intake Review
2. Bottleneck Detection
3. Patient Journey Risk Scoring
4. Scheduling / Coordination Recommendations
5. Communication Drafting (internal notes only)

The synthetic cohort models four branching care pathways (surgery-first,
neoadjuvant therapy, radiation-first, systemic-therapy-only) with correlated
clinical (stage, ECOG, comorbidities, biomarker testing) and operational
(payer type, facility, distance, no-show history) attributes, so delays and
risk emerge from realistic drivers rather than a single linear timeline.
"""

from __future__ import annotations

import random
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuration
RANDOM_SEED = 42
OUTPUT_DIR = Path(__file__).parent
DATA_FILE = OUTPUT_DIR / "synthetic_oncology_scheduling_data.xlsx"
REPORT_FILE = OUTPUT_DIR / "oncology_agent_report.md"
ANALYSIS_FILE = OUTPUT_DIR / "oncology_agent_analysis.xlsx"

TODAY = datetime(2026, 7, 23)  # fixed for reproducibility (matches reference generation time)
# Patient journeys are set in the prior ~12 months: roughly July 2025 → July 2026

COHORT_SIZE = 1200

CANCER_TYPES = [
    "Breast Cancer",
    "Colorectal Cancer",
    "Non-Small Cell Lung Cancer",
    "Prostate Cancer",
    "Melanoma",
    "Pancreatic Cancer",
    "Ovarian Cancer",
    "Head and Neck Cancer",
]

PROVIDERS = [
    "Dr. Elena Carter",
    "Dr. Marcus Rivera",
    "Dr. Priya Patel",
    "Dr. James Okonkwo",
    "Dr. Sarah Nguyen",
]

COORDINATORS = [
    "Nina Alvarez", "Sam Whitfield", "Devon Park", "Rachel O'Brien",
    "Malik Johnson", "Grace Lindqvist",
]

FACILITIES = [
    "Downtown Cancer Center",
    "Northside Oncology Clinic",
    "Riverside Regional Infusion Center",
]

REFERRAL_SOURCES = ["Primary Care", "Emergency Department", "Self-Referral", "Other Specialist"]

FIRST_NAMES = [
    "Jordan", "Alex", "Taylor", "Morgan", "Casey", "Riley", "Quinn",
    "Avery", "Cameron", "Reese", "Parker", "Sawyer", "Harper", "Finley",
]
LAST_NAMES = [
    "Miller", "Brooks", "Chen", "Garcia", "Thompson", "Nguyen", "Patel",
    "Williams", "Khan", "Lopez", "Kim", "Davis", "Rodriguez", "Martinez",
]

STAGES = ["I", "II", "III", "IV"]
ECOG_STATUSES = [0, 1, 2, 3, 4]
PAYERS = ["Commercial", "Medicare", "Medicaid", "Uninsured"]

# Auth outcome weights conditioned on payer type: [Approved, Pending, Denied, Not Required]
AUTH_WEIGHTS_BY_PAYER = {
    "Commercial": [0.65, 0.20, 0.03, 0.12],
    "Medicare": [0.60, 0.22, 0.03, 0.15],
    "Medicaid": [0.40, 0.40, 0.08, 0.12],
    "Uninsured": [0.25, 0.45, 0.15, 0.15],
}

BIOMARKER_APPLICABLE = {"Non-Small Cell Lung Cancer", "Breast Cancer", "Colorectal Cancer"}

TREATMENT_PATHWAYS = ["Surgery-First", "Neoadjuvant Therapy", "Radiation-First", "Systemic Therapy Only"]

# Per-cancer-type pathway weighting: reflects realistic care pattern differences
PATHWAY_WEIGHTS = {
    "Breast Cancer":                  [0.45, 0.40, 0.05, 0.10],
    "Colorectal Cancer":              [0.55, 0.30, 0.05, 0.10],
    "Non-Small Cell Lung Cancer":     [0.25, 0.35, 0.20, 0.20],
    "Prostate Cancer":                [0.55, 0.05, 0.35, 0.05],
    "Melanoma":                       [0.70, 0.10, 0.05, 0.15],
    "Pancreatic Cancer":              [0.20, 0.45, 0.10, 0.25],
    "Ovarian Cancer":                 [0.35, 0.45, 0.05, 0.15],
    "Head and Neck Cancer":           [0.30, 0.10, 0.45, 0.15],
}

HOLIDAY_WINDOWS = [(datetime(2025, 11, 15), datetime(2026, 1, 5))]

# Milestone chain definition per pathway: (field_name, min_days, max_days, gate)
# gate is one of "prereqs" (imaging+pathology+biomarker), "auth" (insurance auth), or None (always allowed)
PATHWAY_STEPS: dict[str, list[tuple[str, int, int, str | None]]] = {
    "Surgery-First": [
        ("tumor_board_date", 7, 21, "prereqs"),
        ("surgical_discussion_date", 5, 20, None),
        ("pre_op_clearance_date", 8, 22, "auth"),
        ("surgery_date", 14, 45, None),
    ],
    "Neoadjuvant Therapy": [
        ("tumor_board_date", 7, 21, "prereqs"),
        ("neoadjuvant_start_date", 10, 25, "auth"),
        ("neoadjuvant_end_date", 60, 120, None),
        ("restaging_imaging_date", 7, 14, None),
        ("surgical_discussion_date", 5, 15, None),
        ("pre_op_clearance_date", 8, 20, None),
        ("surgery_date", 14, 35, None),
    ],
    "Radiation-First": [
        ("tumor_board_date", 7, 21, "prereqs"),
        ("radiation_planning_date", 7, 21, "auth"),
        ("radiation_start_date", 7, 14, None),
        ("radiation_end_date", 28, 42, None),
        ("reassessment_date", 30, 45, None),
    ],
    "Systemic Therapy Only": [
        ("tumor_board_date", 7, 21, "prereqs"),
        ("systemic_therapy_start_date", 10, 21, "auth"),
    ],
}

DEFINITIVE_FIELD_BY_PATHWAY = {
    "Surgery-First": "surgery_date",
    "Neoadjuvant Therapy": "surgery_date",
    "Radiation-First": "radiation_end_date",
    "Systemic Therapy Only": "systemic_therapy_start_date",
}

ALL_MILESTONE_FIELDS = [
    "tumor_board_date", "neoadjuvant_start_date", "neoadjuvant_end_date",
    "restaging_imaging_date", "radiation_planning_date", "radiation_start_date",
    "radiation_end_date", "reassessment_date", "surgical_discussion_date",
    "pre_op_clearance_date", "surgery_date", "systemic_therapy_start_date",
]


# Synthetic data generation
def business_day(d: datetime) -> datetime:
    """Nudge a date off weekends onto the following Monday."""
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d


def in_holiday_window(d: datetime) -> bool:
    return any(start <= d <= end for start, end in HOLIDAY_WINDOWS)


def offset_date(base: datetime, lo: int, hi: int, extra_delay: int = 0) -> datetime:
    """Advance `base` by a random offset, adding seasonal slowdown and any
    accumulated scheduling-friction delay, then snap off weekends."""
    d = base + timedelta(days=random.randint(lo, hi) + extra_delay)
    if in_holiday_window(d):
        d += timedelta(days=random.randint(5, 15))
    return business_day(d)


def random_date(start: datetime, end: datetime) -> datetime:
    delta = end - start
    return start + timedelta(days=random.randint(0, max(delta.days, 0)))


def weighted_choice(options: list, weights: list):
    return random.choices(options, weights=weights)[0]


def gen_stage_ecog_comorbidity(age: int) -> tuple[str, int, int]:
    stage = weighted_choice(STAGES, [0.22, 0.28, 0.30, 0.20])
    stage_idx = STAGES.index(stage)
    ecog_weights = [max(0.05, 0.5 - 0.12 * stage_idx), 0.30, 0.20 + 0.08 * stage_idx,
                     0.05 + 0.05 * stage_idx, 0.02 + 0.03 * stage_idx]
    ecog = weighted_choice(ECOG_STATUSES, ecog_weights)
    comorbidity_lambda = 0.5 + max(0, age - 50) / 25
    comorbidity_count = min(6, sum(1 for _ in range(6) if random.random() < comorbidity_lambda / 6))
    return stage, ecog, comorbidity_count


def gen_pathway(diagnosis: str, stage: str) -> str:
    if stage == "IV" and random.random() < 0.5:
        return "Systemic Therapy Only"
    return weighted_choice(TREATMENT_PATHWAYS, PATHWAY_WEIGHTS[diagnosis])


def priority_from_clinical(stage: str, ecog: int) -> str:
    if stage in ("III", "IV") or ecog >= 2:
        return weighted_choice(["High", "Medium"], [0.75, 0.25])
    return weighted_choice(["High", "Medium", "Low"], [0.15, 0.45, 0.40])


def run_pathway_chain(
    pathway: str,
    npv: datetime,
    prereqs_ok: bool,
    prereq_barrier: str,
    auth_status: str,
    extra_delay: int,
) -> tuple[dict[str, datetime | None], str, str]:
    """Walk a pathway's milestone chain, stopping at the first unmet gate or
    the first milestone still in the future. Returns (dates, status, barrier)."""
    steps = PATHWAY_STEPS[pathway]
    dates: dict[str, datetime | None] = {name: None for name, *_ in steps}
    current = npv
    barrier = ""
    chain_broken = False

    for name, lo, hi, gate in steps:
        if chain_broken:
            continue
        if gate == "prereqs" and not prereqs_ok:
            barrier = prereq_barrier
            chain_broken = True
            continue
        if gate == "auth" and auth_status not in ("Approved", "Not Required"):
            barrier = (
                "Insurance authorization denied – appeal needed" if auth_status == "Denied"
                else "Insurance authorization pending"
            )
            chain_broken = True
            continue
        nd = offset_date(current, lo, hi, extra_delay=extra_delay if name == steps[0][0] else 0)
        dates[name] = nd
        current = nd
        if nd > TODAY:
            chain_broken = True

    definitive_field = DEFINITIVE_FIELD_BY_PATHWAY[pathway]
    definitive_date = dates.get(definitive_field)

    if definitive_date and definitive_date <= TODAY:
        status = "Completed"
        barrier = ""
    elif definitive_date and definitive_date > TODAY:
        status = "Scheduled"
    elif barrier:
        status = "Delayed"
    else:
        status = "In Progress"

    return dates, status, barrier


def generate_synthetic_patients(n: int = COHORT_SIZE) -> pd.DataFrame:
    random.seed(RANDOM_SEED)
    rows = []

    # Pre-defined high-impact scenarios so the report always includes
    # concrete, illustrative cases, layered on top of the larger probabilistic cohort.
    forced_scenarios = {
        0: {  # classic pathology lag + high priority
            "diagnosis": "Colorectal Cancer", "pathway": "Surgery-First",
            "stage": "III", "ecog": 1, "imaging": True, "pathology": False,
            "biomarker": "Pending", "auth": "Pending",
            "referral": datetime(2025, 9, 12),
        },
        1: {  # capacity / long neoadjuvant-to-surgery wait
            "diagnosis": "Breast Cancer", "pathway": "Neoadjuvant Therapy",
            "stage": "II", "ecog": 0, "imaging": True, "pathology": True,
            "biomarker": "Complete", "auth": "Approved",
            "referral": datetime(2025, 8, 5), "extra_delay": 25,
        },
        3: {  # imaging incomplete blocks tumor board
            "diagnosis": "Non-Small Cell Lung Cancer", "pathway": "Neoadjuvant Therapy",
            "stage": "III", "ecog": 1, "imaging": False, "pathology": False,
            "biomarker": "Not Ordered", "auth": "Pending",
            "referral": datetime(2026, 3, 18),
        },
        6: {  # auth pending blocks radiation start
            "diagnosis": "Pancreatic Cancer", "pathway": "Radiation-First",
            "stage": "III", "ecog": 1, "imaging": True, "pathology": True,
            "biomarker": "Not Applicable", "auth": "Pending",
            "referral": datetime(2025, 11, 4),
        },
        8: {  # completed journey (positive control)
            "diagnosis": "Prostate Cancer", "pathway": "Surgery-First",
            "stage": "I", "ecog": 0, "imaging": True, "pathology": True,
            "biomarker": "Not Applicable", "auth": "Approved",
            "referral": datetime(2025, 7, 22),
        },
        11: {  # biomarker turnaround stalls neoadjuvant start
            "diagnosis": "Melanoma", "pathway": "Neoadjuvant Therapy",
            "stage": "III", "ecog": 1, "imaging": True, "pathology": True,
            "biomarker": "Pending", "auth": "Not Required",
            "referral": datetime(2025, 10, 8),
        },
    }

    for i in range(n):
        mrn = f"MRN-{100000 + i * 137}"
        name = f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"
        dob = datetime(random.randint(1948, 1988), random.randint(1, 12), random.randint(1, 28))
        age = TODAY.year - dob.year
        provider = random.choice(PROVIDERS)
        coordinator = random.choice(COORDINATORS)
        facility = random.choice(FACILITIES)
        referral_source = random.choice(REFERRAL_SOURCES)
        payer = weighted_choice(PAYERS, [0.45, 0.25, 0.20, 0.10])
        distance = round(random.uniform(2, 25) if random.random() > 0.15 else random.uniform(25, 90), 1)
        no_show = sum(1 for _ in range(4) if random.random() < (0.10 if payer in ("Medicaid", "Uninsured") else 0.04))
        reschedule = sum(1 for _ in range(4) if random.random() < 0.12)

        forced = forced_scenarios.get(i)
        if forced:
            diagnosis = forced["diagnosis"]
            pathway = forced["pathway"]
            stage = forced["stage"]
            ecog = forced["ecog"]
            _, _, comorbidity_count = gen_stage_ecog_comorbidity(age)
            priority = "High"
            referral = forced["referral"]
            imaging_complete = forced["imaging"]
            pathology_received = forced["pathology"]
            biomarker_status = forced["biomarker"]
            auth_status = forced["auth"]
            extra_delay = forced.get("extra_delay", 0)
        else:
            diagnosis = random.choice(CANCER_TYPES)
            stage, ecog, comorbidity_count = gen_stage_ecog_comorbidity(age)
            pathway = gen_pathway(diagnosis, stage)
            priority = priority_from_clinical(stage, ecog)
            referral = random_date(datetime(2025, 7, 15), datetime(2026, 6, 20))
            imaging_complete = random.choice([True, True, True, False])
            pathology_received = random.choice([True, True, False])
            if diagnosis in BIOMARKER_APPLICABLE:
                biomarker_status = weighted_choice(["Complete", "Pending", "Not Ordered"], [0.55, 0.30, 0.15])
            else:
                biomarker_status = "Not Applicable"
            auth_status = weighted_choice(
                ["Approved", "Pending", "Denied", "Not Required"], AUTH_WEIGHTS_BY_PAYER[payer]
            )
            extra_delay = no_show * random.randint(2, 6) + reschedule * random.randint(1, 4)

        npv = business_day(referral + timedelta(days=random.randint(9, 28)))

        prereqs_ok = imaging_complete and pathology_received and biomarker_status in ("Complete", "Not Applicable")
        if not imaging_complete:
            prereq_barrier = "Imaging not completed"
        elif not pathology_received:
            prereq_barrier = "Pathology not received"
        else:
            prereq_barrier = "Biomarker/genomic testing pending"

        dates, status, barrier = run_pathway_chain(
            pathway, npv, prereqs_ok, prereq_barrier, auth_status, extra_delay
        )

        definitive_field = DEFINITIVE_FIELD_BY_PATHWAY[pathway]
        definitive_date = dates.get(definitive_field)

        row = {
            "mrn": mrn,
            "patient_name": name,
            "dob": dob.strftime("%Y-%m-%d"),
            "age": age,
            "diagnosis_type": diagnosis,
            "stage": stage,
            "ecog_status": ecog,
            "comorbidity_count": comorbidity_count,
            "referral_date": referral.strftime("%Y-%m-%d"),
            "referral_source": referral_source,
            "facility_site": facility,
            "provider": provider,
            "assigned_coordinator": coordinator,
            "payer_type": payer,
            "distance_miles": distance,
            "no_show_count": no_show,
            "reschedule_count": reschedule,
            "new_patient_visit_date": npv.strftime("%Y-%m-%d"),
            "imaging_complete": "Yes" if imaging_complete else "No",
            "pathology_received": "Yes" if pathology_received else "No",
            "biomarker_testing_status": biomarker_status,
            "insurance_auth_status": auth_status,
            "treatment_modality_pathway": pathway,
            "priority_level": priority,
            "scheduling_status": status,
            "barrier_reason": barrier,
            "definitive_treatment_type": definitive_field,
            "definitive_treatment_date": definitive_date.strftime("%Y-%m-%d") if definitive_date else "",
        }
        for field in ALL_MILESTONE_FIELDS:
            d = dates.get(field)
            row[field] = d.strftime("%Y-%m-%d") if d else ""

        rows.append(row)

    df = pd.DataFrame(rows)
    return df


# Agent logic (lightweight sequential "agents")
def parse_date(val: Any) -> datetime | None:
    if pd.isna(val) or val == "" or val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val)[:10], "%Y-%m-%d")
    except Exception:
        return None


def days_between(a: datetime | None, b: datetime | None) -> int | None:
    if a is None or b is None:
        return None
    return (b - a).days


def pathway_milestone_order(pathway: str) -> list[str]:
    return [name for name, *_ in PATHWAY_STEPS[pathway]]


def intake_review_agent(row: pd.Series) -> list[str]:
    """Flag missing prerequisites for progressing the journey."""
    flags = []
    if row["imaging_complete"] != "Yes":
        flags.append("Missing imaging")
    if row["pathology_received"] != "Yes":
        flags.append("Missing pathology report")
    if row["biomarker_testing_status"] == "Pending":
        flags.append("Biomarker testing pending")
    elif row["biomarker_testing_status"] == "Not Ordered":
        flags.append("Biomarker testing not ordered")
    if row["insurance_auth_status"] == "Pending":
        flags.append("Insurance auth pending")
    if row["insurance_auth_status"] == "Denied":
        flags.append("Auth denied – needs appeal")
    if not row["tumor_board_date"] and row["scheduling_status"] != "Completed":
        flags.append("Tumor board review not scheduled")
    pathway = row["treatment_modality_pathway"]
    if pathway in ("Surgery-First", "Neoadjuvant Therapy"):
        if row["tumor_board_date"] and not row["surgical_discussion_date"] and row["scheduling_status"] != "Completed":
            flags.append("Surgical discussion not scheduled")
        if row["surgical_discussion_date"] and not row["pre_op_clearance_date"] and row["scheduling_status"] != "Completed":
            flags.append("Pre-op clearance not scheduled")
    if pathway == "Radiation-First" and row["radiation_planning_date"] and not row["radiation_start_date"] and row["scheduling_status"] != "Completed":
        flags.append("Radiation start not scheduled")
    if row["no_show_count"] >= 2:
        flags.append("Elevated no-show history")
    if not flags and row["scheduling_status"] == "Delayed":
        flags.append("Capacity / slot availability issue")
    return flags


def bottleneck_detection_agent(row: pd.Series) -> dict[str, Any]:
    """Calculate key interval lengths and identify primary bottleneck."""
    referral = parse_date(row["referral_date"])
    npv = parse_date(row["new_patient_visit_date"])
    pathway = row["treatment_modality_pathway"]
    definitive_date = parse_date(row["definitive_treatment_date"])

    intervals: dict[str, Any] = {
        "referral_to_npv_days": days_between(referral, npv),
        "referral_to_definitive_days": days_between(referral, definitive_date),
    }
    # First populated milestone after NPV, for a generic "npv to next step" measure
    milestone_fields = pathway_milestone_order(pathway)
    first_milestone_date = None
    for f in milestone_fields:
        d = parse_date(row[f])
        if d:
            first_milestone_date = d
            break
    intervals["npv_to_next_step_days"] = days_between(npv, first_milestone_date)

    if intervals["referral_to_definitive_days"] is None and referral:
        last = definitive_date
        if last is None:
            for f in reversed(milestone_fields):
                d = parse_date(row[f])
                if d:
                    last = d
                    break
        last = last or npv or referral
        base_remaining = {"Surgery-First": 40, "Neoadjuvant Therapy": 70, "Radiation-First": 55,
                           "Systemic Therapy Only": 30}.get(pathway, 45)
        remaining = base_remaining + (20 if row.get("scheduling_status") == "Delayed" else 0)
        intervals["projected_total_days"] = (last - referral).days + remaining
    else:
        intervals["projected_total_days"] = intervals["referral_to_definitive_days"]

    # Primary bottleneck logic, in clinical/operational priority order
    primary = "None identified"
    secondary = ""
    if row["imaging_complete"] != "Yes":
        primary = "Imaging incomplete"
        secondary = "Staging / planning incomplete"
    elif row["pathology_received"] != "Yes":
        primary = "Pathology not received"
        secondary = "Blocks tumor board review and treatment planning"
    elif row["biomarker_testing_status"] in ("Pending", "Not Ordered"):
        primary = "Biomarker/genomic testing not resolved"
        secondary = "Blocks treatment selection (targeted therapy eligibility)"
    elif row["insurance_auth_status"] == "Pending":
        primary = "Insurance authorization pending"
        secondary = "Prevents firm treatment start booking"
    elif row["insurance_auth_status"] == "Denied":
        primary = "Insurance authorization denied"
        secondary = "Requires appeal / peer-to-peer review"
    elif not row["tumor_board_date"]:
        primary = "Tumor board review not scheduled"
        secondary = "Multidisciplinary conference capacity or intake lag"
    elif pathway in ("Surgery-First", "Neoadjuvant Therapy") and not row["surgical_discussion_date"]:
        primary = "Surgical discussion not scheduled"
        secondary = "Provider template / coordinator bandwidth"
    elif pathway == "Radiation-First" and not row["radiation_start_date"] and row["radiation_planning_date"]:
        primary = "Radiation start delayed after planning"
        secondary = "Linac / treatment slot capacity"
    elif row["no_show_count"] >= 3:
        primary = "Repeated no-shows disrupting scheduling"
        secondary = "Consider transportation support / reminder outreach"
    elif intervals.get("projected_total_days") and intervals["projected_total_days"] > 150:
        primary = "Long overall projected diagnosis-to-treatment interval"
        secondary = "Review end-to-end pathway capacity for this diagnosis/pathway"
    elif row["scheduling_status"] == "Delayed":
        primary = "General capacity or coordination lag"
        secondary = "Review open slots and cancellation list"

    intervals["primary_bottleneck"] = primary
    intervals["secondary_note"] = secondary
    return intervals


def risk_scoring_agent(row: pd.Series, intervals: dict) -> dict[str, Any]:
    """Assign delay risk score and level."""
    projected = intervals.get("projected_total_days") or intervals.get("referral_to_definitive_days") or 0
    flags = intake_review_agent(row)

    score = 0
    # Time component
    if projected >= 150:
        score += 35
    elif projected >= 120:
        score += 25
    elif projected >= 90:
        score += 15
    elif projected >= 60:
        score += 5

    # Missing critical items
    if "Missing pathology report" in flags:
        score += 20
    if "Biomarker testing pending" in flags or "Biomarker testing not ordered" in flags:
        score += 10
    if "Insurance auth pending" in flags or "Auth denied – needs appeal" in flags:
        score += 15
    if "Missing imaging" in flags:
        score += 15
    if "Tumor board review not scheduled" in flags:
        score += 10
    if "Elevated no-show history" in flags:
        score += 8

    # Clinical acuity
    if row["stage"] in ("III", "IV"):
        score += 8
    if row["ecog_status"] >= 2:
        score += 8
    score += min(10, row["comorbidity_count"] * 2)

    # Operational / access
    if row["payer_type"] in ("Medicaid", "Uninsured"):
        score += 6

    # Priority boost
    if row["priority_level"] == "High":
        score += 8
    elif row["priority_level"] == "Medium":
        score += 4

    # Status
    if row["scheduling_status"] == "Delayed":
        score += 10
    if row["scheduling_status"] == "Completed":
        score = max(0, score - 30)

    score = min(100, score)

    if score >= 70:
        level = "High"
    elif score >= 40:
        level = "Moderate"
    else:
        level = "Low"

    return {"risk_score": score, "risk_level": level, "projected_days": projected}


def recommendation_agent(row: pd.Series, intervals: dict, risk: dict) -> str:
    """Produce a concrete next-action recommendation."""
    primary = intervals["primary_bottleneck"]
    level = risk["risk_level"]

    if row["scheduling_status"] == "Completed":
        return "Journey complete – archive / outcomes tracking only."

    if "Pathology" in primary:
        return (
            "High priority: Contact outside facility / pathology lab to expedite report. "
            "Hold firm treatment start until received. Escalate if >7 days outstanding."
        )
    if "Biomarker" in primary:
        return (
            "Follow up with molecular pathology / reference lab on turnaround time. "
            "If not yet ordered, place order today — this gates treatment selection."
        )
    if "authorization denied" in primary.lower():
        return "Initiate formal appeal and request peer-to-peer review with medical director today."
    if "authorization" in primary.lower():
        return (
            "Follow up with insurance / financial counselor today. "
            "If pending >10 days, escalate to authorization specialist and consider peer-to-peer."
        )
    if "Imaging" in primary:
        return "Coordinate imaging completion; place order if missing and schedule ASAP."
    if "Tumor board" in primary:
        return "Add case to next available multidisciplinary tumor board slot; confirm required data is packaged in advance."
    if "Surgical discussion" in primary:
        return (
            "Review surgical discussion template availability for the assigned provider. "
            "Consider converting under-utilized return slots or using cancellation list for high-priority cases."
        )
    if "Radiation start delayed" in primary:
        return "Check linac scheduling capacity and simulation slot availability; escalate to radiation oncology charge therapist."
    if "no-shows" in primary:
        return f"Coordinate with {row['assigned_coordinator']} on transportation support and proactive reminder calls; consider consolidating visits."
    if "overall projected" in primary:
        return "Escalate to service line leadership for capacity review across this diagnosis/pathway; consider cross-site scheduling."
    if level == "High":
        return (
            "Prioritize for daily huddle. Assign coordinator owner and set 48-hour follow-up. "
            "Evaluate whether a cancellation slot or template conversion can pull the case forward."
        )
    return (
        "Continue standard coordination. Re-check status in 5–7 days. "
        "Monitor for new barriers (auth, pathology, biomarker, clearance)."
    )


def communication_draft_agent(row: pd.Series, intervals: dict, risk: dict, recommendation: str) -> str:
    """Draft an internal note only (never auto-send to patient)."""
    return (
        f"Internal note – {row['mrn']} ({row['patient_name']})\n"
        f"Diagnosis: {row['diagnosis_type']} | Stage: {row['stage']} | Pathway: {row['treatment_modality_pathway']}\n"
        f"Priority: {row['priority_level']} | Risk: {risk['risk_level']} ({risk['risk_score']}) | Payer: {row['payer_type']}\n"
        f"Current status: {row['scheduling_status']} | Coordinator: {row['assigned_coordinator']}\n"
        f"Primary bottleneck: {intervals['primary_bottleneck']}\n"
        f"Projected / actual diagnosis-to-treatment: ~{risk['projected_days']} days\n"
        f"Recommended action: {recommendation}\n"
        f"— Generated by Synthetic Oncology Scheduling Agent (reference implementation)"
    )


# Orchestration and reporting
def run_full_analysis(df: pd.DataFrame) -> pd.DataFrame:
    results = []
    for _, row in df.iterrows():
        flags = intake_review_agent(row)
        intervals = bottleneck_detection_agent(row)
        risk = risk_scoring_agent(row, intervals)
        rec = recommendation_agent(row, intervals, risk)
        draft = communication_draft_agent(row, intervals, risk, rec)

        results.append({
            **row.to_dict(),
            "missing_items": "; ".join(flags) if flags else "None",
            "referral_to_npv_days": intervals["referral_to_npv_days"],
            "npv_to_next_step_days": intervals["npv_to_next_step_days"],
            "referral_to_definitive_days": intervals["referral_to_definitive_days"],
            "projected_total_days": intervals["projected_total_days"],
            "primary_bottleneck": intervals["primary_bottleneck"],
            "secondary_note": intervals["secondary_note"],
            "risk_score": risk["risk_score"],
            "risk_level": risk["risk_level"],
            "recommended_action": rec,
            "internal_draft_note": draft,
        })
    return pd.DataFrame(results)


def style_excel(ws, header_fill: str = "1F4E79"):
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=header_fill)
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def build_cohort_summary(analysis_df: pd.DataFrame) -> pd.DataFrame:
    frames = []
    for dim in ["treatment_modality_pathway", "stage", "payer_type", "facility_site"]:
        g = analysis_df.groupby(dim).agg(
            patient_count=("mrn", "count"),
            avg_risk_score=("risk_score", "mean"),
            delayed_count=("scheduling_status", lambda s: (s == "Delayed").sum()),
            high_risk_count=("risk_level", lambda s: (s == "High").sum()),
        ).reset_index().rename(columns={dim: "value"})
        g.insert(0, "dimension", dim)
        g["avg_risk_score"] = g["avg_risk_score"].round(1)
        frames.append(g)
    return pd.concat(frames, ignore_index=True)


def save_analysis_excel(raw_df: pd.DataFrame, analysis_df: pd.DataFrame):
    wb = Workbook()

    # Sheet 1: Raw synthetic data
    ws1 = wb.active
    ws1.title = "Synthetic Data"
    for r in dataframe_to_rows(raw_df, index=False, header=True):
        ws1.append(r)
    style_excel(ws1)

    # Sheet 2: Full agent output
    ws2 = wb.create_sheet("Agent Analysis")
    cols = [
        "mrn", "patient_name", "diagnosis_type", "stage", "ecog_status",
        "treatment_modality_pathway", "priority_level",
        "scheduling_status", "risk_level", "risk_score", "projected_total_days",
        "primary_bottleneck", "missing_items", "recommended_action",
        "payer_type", "facility_site", "assigned_coordinator",
        "referral_date", "new_patient_visit_date", "definitive_treatment_type",
        "definitive_treatment_date", "provider", "barrier_reason",
    ]
    view = analysis_df[cols]
    for r in dataframe_to_rows(view, index=False, header=True):
        ws2.append(r)
    style_excel(ws2, header_fill="C65911")

    # Sheet 3: High-risk only
    ws3 = wb.create_sheet("High Risk Patients")
    high = analysis_df[analysis_df["risk_level"] == "High"][cols]
    for r in dataframe_to_rows(high, index=False, header=True):
        ws3.append(r)
    style_excel(ws3, header_fill="C00000")

    # Sheet 4: Cohort summary (aggregate breakdowns for a cohort too large to eyeball row by row)
    ws4 = wb.create_sheet("Cohort Summary")
    summary_df = build_cohort_summary(analysis_df)
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws4.append(r)
    style_excel(ws4, header_fill="375623")

    # Column widths (approximate)
    for ws in (ws1, ws2, ws3, ws4):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, min(len(str(cell.value)), 45))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(ANALYSIS_FILE)
    print(f"Saved detailed analysis → {ANALYSIS_FILE}")


def generate_markdown_report(analysis_df: pd.DataFrame) -> str:
    high = analysis_df[analysis_df["risk_level"] == "High"].sort_values("risk_score", ascending=False)
    moderate = analysis_df[analysis_df["risk_level"] == "Moderate"]
    delayed = analysis_df[analysis_df["scheduling_status"] == "Delayed"]

    TOP_N_DETAILED = 25

    lines = [
        "# Synthetic Oncology Scheduling Optimization Agent – Reference Report",
        "",
        f"**Generated:** {TODAY.strftime('%Y-%m-%d')}  ",
        f"**Records analyzed:** {len(analysis_df)} synthetic patients  ",
        f"**High-risk patients:** {len(high)}  ",
        f"**Moderate-risk patients:** {len(moderate)}  ",
        f"**Currently delayed:** {len(delayed)}  ",
        "",
        "---",
        "",
        "## Executive Snapshot",
        "",
        "This report uses **100% synthetic data** (fake MRNs, names, and dates) across a "
        f"{len(analysis_df)}-patient cohort spanning four branching care pathways "
        "(surgery-first, neoadjuvant therapy, radiation-first, systemic-therapy-only). "
        "It illustrates how an operational AI copilot could help oncology scheduling teams "
        "surface delays, missing documents, and capacity issues earlier — without making clinical decisions "
        "or contacting patients autonomously.",
        "",
        "### Key Findings",
        f"- {len(high)} patients scored **High** risk (projected/actual diagnosis-to-treatment interval, "
        "missing critical items, clinical acuity, and access barriers).",
        "- Most common primary bottlenecks observed in this synthetic cohort:",
    ]

    bottleneck_counts = analysis_df["primary_bottleneck"].value_counts().head(8)
    for b, c in bottleneck_counts.items():
        lines.append(f"  - {b}: {c} patients")

    lines += ["", "### Cohort Composition"]
    for label, col in [("Treatment pathway", "treatment_modality_pathway"), ("Stage", "stage"), ("Payer type", "payer_type")]:
        lines.append(f"- **{label}:** " + ", ".join(f"{k} ({v})" for k, v in analysis_df[col].value_counts().items()))

    lines += [
        "",
        "---",
        "",
        f"## High-Risk Patients (Action List, top {min(TOP_N_DETAILED, len(high))} by risk score)",
        "",
    ]

    if high.empty:
        lines.append("_No high-risk patients in this synthetic run._")
    else:
        for _, r in high.head(TOP_N_DETAILED).iterrows():
            lines += [
                f"### {r['mrn']} — {r['patient_name']}",
                f"- **Diagnosis / Stage / Pathway:** {r['diagnosis_type']} / {r['stage']} / {r['treatment_modality_pathway']}  ",
                f"- **Priority / Risk:** {r['priority_level']} / **{r['risk_level']}** (score {r['risk_score']})  ",
                f"- **Payer / Facility / Coordinator:** {r['payer_type']} / {r['facility_site']} / {r['assigned_coordinator']}  ",
                f"- **Projected days (referral → definitive treatment):** ~{r['projected_total_days']}  ",
                f"- **Primary bottleneck:** {r['primary_bottleneck']}  ",
                f"- **Missing / flags:** {r['missing_items']}  ",
                f"- **Recommended action:** {r['recommended_action']}  ",
                "",
                "<details><summary>Internal draft note (copy-paste ready)</summary>",
                "",
                "```",
                r["internal_draft_note"],
                "```",
                "</details>",
                "",
            ]
        if len(high) > TOP_N_DETAILED:
            lines.append(
                f"_{len(high) - TOP_N_DETAILED} additional high-risk patients omitted for brevity — "
                f"see the \"High Risk Patients\" sheet in `{ANALYSIS_FILE.name}` for the full list._"
            )
            lines.append("")

    lines += [
        "---",
        "",
        "## Cohort Summary (Aggregate)",
        "",
        "| Dimension | Value | Patients | Avg Risk | Delayed | High Risk |",
        "|---|---|---|---|---|---|",
    ]
    summary_df = build_cohort_summary(analysis_df)
    for _, r in summary_df.iterrows():
        lines.append(
            f"| {r['dimension']} | {r['value']} | {r['patient_count']} | {r['avg_risk_score']} | "
            f"{r['delayed_count']} | {r['high_risk_count']} |"
        )

    lines += [
        "",
        "---",
        "",
        "## How the Agents Work",
        "",
        "1. **Intake Review Agent** – Checks imaging, pathology, biomarker testing, auth, and whether key "
        "visits/reviews are scheduled. Surfaces missing prerequisites before the case stalls further.",
        "2. **Bottleneck Detection Agent** – Walks each patient's pathway-specific milestone chain "
        "(surgery-first, neoadjuvant, radiation-first, or systemic-only) and names the dominant delay driver.",
        "3. **Patient Journey Risk Agent** – Combines timeline length, missing critical items, clinical "
        "acuity (stage, ECOG, comorbidities), and access factors (payer, no-shows) into a 0–100 score.",
        "4. **Scheduling Optimization / Recommendation Agent** – Suggests concrete next steps (expedite "
        "pathology/biomarker turnaround, review cancellation list, convert slots, escalate auth appeals, etc.).",
        "5. **Communication Drafting Agent** – Produces an internal coordinator note only. Never auto-contacts "
        "patients or makes clinical decisions.",
        "",
        "**Design philosophy:** AI as an operational copilot for the front-end scheduling and coordination "
        "team — not an autonomous scheduler.",
        "",
        "---",
        "",
        "## Files Produced",
        "",
        f"- `{DATA_FILE.name}` – clean synthetic source data",
        f"- `{ANALYSIS_FILE.name}` – four sheets (raw, full analysis, high-risk only, cohort summary)",
        f"- `{REPORT_FILE.name}` – this summary",
        "",
        "All data in this report is fully synthetic; no real patient, provider, payer, or facility "
        "information is used or represented.",
        "",
    ]
    return "\n".join(lines)


def main():
    print("=" * 60)
    print("Synthetic Oncology Scheduling Optimization Agent")
    print("=" * 60)

    # 1. Generate / load synthetic data
    print("\n[1] Generating synthetic patient cohort...")
    raw_df = generate_synthetic_patients(n=COHORT_SIZE)
    raw_df.to_excel(DATA_FILE, index=False)
    print(f"    → {DATA_FILE} ({len(raw_df)} patients)")

    # 2. Run agents
    print("\n[2] Running multi-agent style analysis...")
    analysis_df = run_full_analysis(raw_df)

    # 3. Persist
    print("\n[3] Writing Excel analysis workbook...")
    save_analysis_excel(raw_df, analysis_df)

    print("\n[4] Writing Markdown report...")
    md = generate_markdown_report(analysis_df)
    REPORT_FILE.write_text(md, encoding="utf-8")
    print(f"    → {REPORT_FILE}")

    # 4. Console dashboard
    print("\n" + "=" * 60)
    print("QUICK DASHBOARD (High-Risk First)")
    print("=" * 60)

    high = analysis_df[analysis_df["risk_level"] == "High"].sort_values("risk_score", ascending=False)
    if high.empty:
        print("No high-risk patients this run.")
    else:
        for _, r in high.head(6).iterrows():
            print(f"\n• {r['mrn']} | {r['patient_name']} | {r['diagnosis_type']} | {r['treatment_modality_pathway']}")
            print(f"  Risk: {r['risk_level']} ({r['risk_score']}) | Status: {r['scheduling_status']}")
            print(f"  Bottleneck: {r['primary_bottleneck']}")
            print(f"  Action: {r['recommended_action'][:110]}...")

    print("\n" + "-" * 60)
    print("Summary counts:")
    print(analysis_df["risk_level"].value_counts().to_string())
    print("\nTop bottlenecks:")
    print(analysis_df["primary_bottleneck"].value_counts().head(8).to_string())
    print("\nDone. Open the Excel and Markdown files for the full output package.")
    print("=" * 60)


if __name__ == "__main__":
    main()
