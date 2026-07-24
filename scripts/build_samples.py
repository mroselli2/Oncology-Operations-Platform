#!/usr/bin/env python3
"""Builds the curated public sample workbook and report from the published
snapshot: samples/oncology_operations_sample.xlsx and
samples/oncology_operations_sample_report.md.
"""

from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from oncology_pipeline.reports.excel_report import style_excel

REPO_ROOT = Path(__file__).parent.parent
SNAPSHOT_PATH = REPO_ROOT / "published_data" / "oncology_operations_snapshot.db"
WORKBOOK_PATH = REPO_ROOT / "samples" / "oncology_operations_sample.xlsx"
REPORT_PATH = REPO_ROOT / "samples" / "oncology_operations_sample_report.md"

HUDDLE_TOP_N = 30
HIGH_RISK_TOP_N = 50

DATA_DICTIONARY = [
    ("mrn", "Synthetic patient identifier (ONC-MRN-######)"),
    ("diagnosis_type", "Cancer type"),
    ("stage", "Clinical stage (I-IV)"),
    ("treatment_modality_pathway", "Care pathway (Surgery-First, Neoadjuvant Therapy, Radiation-First, Systemic Therapy Only)"),
    ("risk_score", "0-100 deterministic delay-risk score"),
    ("risk_level", "High / Moderate / Low, derived from risk_score"),
    ("primary_bottleneck", "Dominant delay driver identified by the bottleneck-detection agent"),
    ("scheduling_status", "Delayed / Scheduled / In Progress / Completed"),
    ("insurance_auth_status", "Approved / Pending / Denied / Not Required"),
    ("recommended_action", "Deterministic next operational action"),
    ("candidate_facility_id / candidate_date", "Ranked appointment slot candidate and its date"),
    ("reason_code", "Why a slot candidate ranked where it did (e.g. IN_NETWORK_PREFERRED, NEXT_AVAILABLE)"),
    ("weekly_capacity / avg_lead_time_days", "Facility service capacity and typical scheduling lead time"),
]


def load_analysis_df(conn: sqlite3.Connection) -> pd.DataFrame:
    return pd.read_sql_query(
        """
        SELECT p.*, ao.risk_score, ao.risk_level, ao.primary_bottleneck, ao.missing_items,
               ao.recommended_action, ao.projected_total_days,
               coo.full_name AS coordinator_name, fac.name AS facility_name,
               pay.payer_name AS payer_name
        FROM patients p
        LEFT JOIN agent_outputs ao ON ao.patient_id = p.patient_id
        LEFT JOIN coordinators coo ON coo.coordinator_id = p.assigned_coordinator_id
        LEFT JOIN facilities fac ON fac.facility_id = p.facility_id
        LEFT JOIN payers pay ON pay.payer_id = p.payer_id
        """,
        conn,
    )


def build_cohort_summary(df: pd.DataFrame) -> pd.DataFrame:
    frames = []
    for dim in ["treatment_modality_pathway", "stage", "payer_type", "facility_name"]:
        g = df.groupby(dim).agg(
            patient_count=("mrn", "count"),
            avg_risk_score=("risk_score", "mean"),
            delayed_count=("scheduling_status", lambda s: (s == "Delayed").sum()),
            high_risk_count=("risk_level", lambda s: (s == "High").sum()),
        ).reset_index().rename(columns={dim: "value"})
        g.insert(0, "dimension", dim)
        g["avg_risk_score"] = g["avg_risk_score"].round(1)
        frames.append(g)
    return pd.concat(frames, ignore_index=True)


def build_workbook(conn: sqlite3.Connection, df: pd.DataFrame) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Cohort Summary"
    for r in dataframe_to_rows(build_cohort_summary(df), index=False, header=True):
        ws1.append(r)
    style_excel(ws1, header_fill="1F4E79")

    huddle_cols = ["mrn", "diagnosis_type", "treatment_modality_pathway", "risk_level",
                   "risk_score", "primary_bottleneck", "recommended_action", "coordinator_name"]
    huddle = df[df["scheduling_status"] != "Completed"].sort_values("risk_score", ascending=False)[huddle_cols].head(HUDDLE_TOP_N)
    ws2 = wb.create_sheet("Daily Huddle Queue")
    for r in dataframe_to_rows(huddle, index=False, header=True):
        ws2.append(r)
    style_excel(ws2, header_fill="C65911")

    high_risk_cols = ["mrn", "diagnosis_type", "stage", "treatment_modality_pathway",
                       "risk_score", "primary_bottleneck", "insurance_auth_status",
                       "recommended_action", "facility_name", "payer_name"]
    high_risk = df[df["risk_level"] == "High"].sort_values("risk_score", ascending=False)[high_risk_cols].head(HIGH_RISK_TOP_N)
    ws3 = wb.create_sheet("High-Risk Cases")
    for r in dataframe_to_rows(high_risk, index=False, header=True):
        ws3.append(r)
    style_excel(ws3, header_fill="C00000")

    slots = pd.read_sql_query(
        """
        SELECT s.patient_id, s.appointment_type_code, s.rank, f.name AS candidate_facility,
               s.candidate_date, s.network_status, s.distance_miles, s.reason_code
        FROM slot_recommendations s
        JOIN facilities f ON f.facility_id = s.candidate_facility_id
        ORDER BY s.patient_id, s.rank
        """,
        conn,
    )
    ws4 = wb.create_sheet("Slot Recommendations")
    for r in dataframe_to_rows(slots, index=False, header=True):
        ws4.append(r)
    style_excel(ws4, header_fill="375623")

    facility_summary = pd.read_sql_query(
        """
        SELECT f.name AS facility_name, f.facility_type, fc.service_type,
               fc.weekly_capacity, fc.avg_lead_time_days
        FROM facility_capabilities fc JOIN facilities f ON f.facility_id = fc.facility_id
        ORDER BY f.name, fc.service_type
        """,
        conn,
    )
    ws5 = wb.create_sheet("Facility Summary")
    for r in dataframe_to_rows(facility_summary, index=False, header=True):
        ws5.append(r)
    style_excel(ws5, header_fill="203864")

    payer_auth = pd.read_sql_query(
        """
        SELECT pay.payer_name, pay.payer_type, au.auth_status, COUNT(*) AS request_count,
               ROUND(AVG(au.total_time_spent_on_phone_min), 1) AS avg_phone_minutes
        FROM authorizations au JOIN payers pay ON pay.payer_id = au.payer_id
        GROUP BY pay.payer_name, au.auth_status ORDER BY pay.payer_name, au.auth_status
        """,
        conn,
    )
    ws6 = wb.create_sheet("Payer Authorization")
    for r in dataframe_to_rows(payer_auth, index=False, header=True):
        ws6.append(r)
    style_excel(ws6, header_fill="7F6000")

    ws7 = wb.create_sheet("Data Dictionary")
    dd_df = pd.DataFrame(DATA_DICTIONARY, columns=["field", "description"])
    for r in dataframe_to_rows(dd_df, index=False, header=True):
        ws7.append(r)
    style_excel(ws7, header_fill="404040")

    for ws in (ws1, ws2, ws3, ws4, ws5, ws6, ws7):
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK_PATH)


def build_report(df: pd.DataFrame) -> str:
    high = df[df["risk_level"] == "High"]
    delayed = df[df["scheduling_status"] == "Delayed"]
    top_bottlenecks = df["primary_bottleneck"].value_counts().head(6)

    lines = [
        "# Oncology Operations Sample Report",
        "",
        "A curated sample of the deterministic operational analysis produced by the "
        "oncology operations pipeline, generated from the published baseline snapshot "
        "(scenario: baseline, seed: 42, 1,200 synthetic patients).",
        "",
        "All data is fully synthetic. No real patient, provider, payer, or facility "
        "information is used or represented.",
        "",
        "## Cohort Snapshot",
        "",
        f"- Patients analyzed: {len(df)}",
        f"- High risk: {len(high)}  |  Delayed: {len(delayed)}",
        "",
        "## Most Common Bottlenecks",
        "",
    ]
    for bottleneck, count in top_bottlenecks.items():
        lines.append(f"- {bottleneck}: {count} patients")

    lines += [
        "",
        "## Pathway Distribution",
        "",
    ]
    for pathway, count in df["treatment_modality_pathway"].value_counts().items():
        lines.append(f"- {pathway}: {count} patients")

    lines += [
        "",
        "## How to Read This Sample",
        "",
        "- **Cohort Summary** -- aggregate counts and average risk by pathway, stage, payer type, and facility.",
        "- **Daily Huddle Queue** -- the top active cases ranked by risk score, in the same "
        "shape the operational dashboard presents them.",
        "- **High-Risk Cases** -- detail on the highest-risk patients, including the "
        "deterministic bottleneck and recommended action for each.",
        "- **Slot Recommendations** -- ranked candidate appointment slots with the reason "
        "each ranking was assigned.",
        "- **Facility Summary** -- service capacity and typical lead time by facility.",
        "- **Payer Authorization** -- authorization volume and average phone time by payer and status.",
        "- **Data Dictionary** -- field definitions for the sheets above.",
        "",
        "## Scope",
        "",
        "This sample reflects the deterministic layer only: risk scores, bottleneck "
        "classifications, and slot recommendations are computed in Python from the "
        "synthetic dataset. The platform also supports an optional, bounded LLM-assisted "
        "narrative layer for a small subset of cases; that layer is not represented in "
        "this static sample.",
        "",
    ]
    return "\n".join(lines)


def main() -> None:
    conn = sqlite3.connect(SNAPSHOT_PATH)
    df = load_analysis_df(conn)
    build_workbook(conn, df)
    REPORT_PATH.write_text(build_report(df), encoding="utf-8")
    conn.close()
    print(f"Wrote {WORKBOOK_PATH}")
    print(f"Wrote {REPORT_PATH}")


if __name__ == "__main__":
    main()
