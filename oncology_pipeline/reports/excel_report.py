"""Excel workbook: agent analysis, high-risk patients, cohort summary,
facility network, and a data-quality sheet."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

ANALYSIS_COLS = [
    "mrn", "first_name", "last_name", "diagnosis_type", "stage", "ecog_status",
    "treatment_modality_pathway", "priority_level",
    "scheduling_status", "risk_level", "risk_score", "projected_total_days",
    "primary_bottleneck", "missing_items", "recommended_action",
    "payer_type", "payer_name", "facility_name", "assigned_coordinator_name",
    "referral_date", "new_patient_visit_date", "definitive_treatment_type",
    "definitive_treatment_date", "assigned_provider_name", "barrier_reason",
]


def style_excel(ws, header_fill: str = "1F4E79"):
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=header_fill)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, min(len(str(cell.value)), 45))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2


def build_cohort_summary(analysis_df: pd.DataFrame) -> pd.DataFrame:
    frames = []
    for dim in ["treatment_modality_pathway", "stage", "payer_type", "facility_name"]:
        g = analysis_df.groupby(dim).agg(
            patient_count=("mrn", "count"),
            avg_risk_score=("risk_score", "mean"),
            delayed_count=("scheduling_status", lambda s: (s == "Delayed").sum()),
            high_risk_count=("risk_level", lambda s: (s == "High").sum()),
        ).reset_index().rename(columns={dim: "value"})
        g.insert(0, "dimension", dim)
        g["avg_risk_score"] = g["avg_risk_score"].round(1)
        frames.append(g)
    return pd.concat(frames, ignore_index=True)


def build_data_quality(conn: sqlite3.Connection, analysis_df: pd.DataFrame) -> pd.DataFrame:
    checks = []
    checks.append(("patients missing primary_dx_code (real taxonomy unmapped)",
                    int(analysis_df["primary_dx_code"].isna().sum())))
    checks.append(("patients with non-ONC id anywhere (mrn)",
                    int((~analysis_df["mrn"].astype(str).str.startswith("ONC-")).sum())))
    checks.append(("patients missing payer (Uninsured, expected)",
                    int(analysis_df["payer_id"].isna().sum())))
    checks.append(("patients with zero missing_items", int((analysis_df["missing_items"] == "None").sum())))
    total = len(analysis_df)
    checks.append(("total patients", total))
    return pd.DataFrame(checks, columns=["check", "value"])


def save_analysis_excel(conn: sqlite3.Connection, analysis_df: pd.DataFrame, output_path: Path):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Agent Analysis"
    view = analysis_df[ANALYSIS_COLS]
    for r in dataframe_to_rows(view, index=False, header=True):
        ws1.append(r)
    style_excel(ws1, header_fill="C65911")

    ws2 = wb.create_sheet("High Risk Patients")
    high = analysis_df[analysis_df["risk_level"] == "High"][ANALYSIS_COLS]
    for r in dataframe_to_rows(high, index=False, header=True):
        ws2.append(r)
    style_excel(ws2, header_fill="C00000")

    ws3 = wb.create_sheet("Cohort Summary")
    summary_df = build_cohort_summary(analysis_df)
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws3.append(r)
    style_excel(ws3, header_fill="375623")

    ws4 = wb.create_sheet("Facility Network")
    network_df = pd.read_sql_query(
        """SELECT f.name AS facility_name, f.facility_type, fc.service_type,
                  fc.weekly_capacity, fc.avg_lead_time_days
           FROM facility_capabilities fc JOIN facilities f ON f.facility_id = fc.facility_id
           ORDER BY f.name, fc.service_type""",
        conn,
    )
    for r in dataframe_to_rows(network_df, index=False, header=True):
        ws4.append(r)
    style_excel(ws4, header_fill="203864")

    ws5 = wb.create_sheet("Data Quality")
    dq_df = build_data_quality(conn, analysis_df)
    for r in dataframe_to_rows(dq_df, index=False, header=True):
        ws5.append(r)
    style_excel(ws5, header_fill="7F6000")

    for ws in (ws1, ws2, ws3, ws4, ws5):
        _autosize(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
